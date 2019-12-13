import os
import time

import pandas as pd
import pymysql
import numpy as np

from cryptocompy import price
from openpyxl import load_workbook, Workbook


class MySQL:
    def __init__(self, item_id):
        # Connect to MySQL
        self._conn = pymysql.connect(
            host='localhost',  # mysql server address
            port=3306,  # port num
            user='root',  # username
            passwd='root',  # password
            db='posts',
            charset='utf8mb4',
        )
        self._cur = self._conn.cursor()
        self._table = item_id

    def __del__(self):
        self._conn.close()

    def searchTweetsByHour(self):
        sql = 'SELECT time, sentiment FROM ' + self._table + \
            ' WHERE time > DATE_SUB(NOW(), INTERVAL 60 MINUTE);'
        df = pd.read_sql(sql, con=self._conn)
        amount_hour = len(df)
        sentiment_hour = df.mean().sentiment

        sql = 'select time, sentiment from ' + self._table
        df = pd.read_sql(sql, con=self._conn)
        amount_all = len(df)

        sql = 'SELECT time FROM ' + self._table + ' limit 1;'
        time_earlist = pd.read_sql(sql, con=self._conn).time[0]
        sql = 'SELECT time from ' + self._table + \
            ' where time = (SELECT max(time) FROM ' + self._table + ')'
        time_latest = pd.read_sql(sql, con=self._conn).time[0]

        time_earlist = pd.to_datetime(time_earlist, format='%Y-%m-%d %H:%M:%S')
        time_latest = pd.to_datetime(time_latest, format='%Y-%m-%d %H:%M:%S')

        amount_avg = amount_all / \
            ((time_latest - time_earlist).total_seconds() / 3600)

        sentiment_avg = df.mean().sentiment

        amount_spike = (amount_hour - amount_avg) / amount_avg
        sentiment_spike = (sentiment_hour - sentiment_avg) / sentiment_avg

        result = {
            'amount_hour': amount_hour,
            'amount_avg': amount_avg,
            'amount_spike': amount_spike,
            'sentiment_hour': sentiment_hour,
            'sentiment_avg': sentiment_avg,
            'sentiment_spike': sentiment_spike
        }
        return result


def gen_crypto_data(item_id):
    result = MySQL(item_id).searchTweetsByHour()
    fsym = item_id.upper()

    price_data = price.get_current_price(fsym,
                                         'USD',
                                         e='all',
                                         try_conversion=True,
                                         full=True,
                                         format='raw')[fsym]['USD']
    result['price'] = price_data['PRICE']
    result['coin_supply'] = price_data['SUPPLY']
    result['market_cap'] = price_data['MKTCAP']
    result['velocity'] = result['price'] * \
        result['coin_supply'] / result['market_cap']

    result['high'] = 0
    result['low'] = float("inf")
    result['volume_sold'] = 0
    for data in price.get_historical_data(fsym,
                                          'USD',
                                          'minute',
                                          aggregate=10,
                                          limit=6):
        result['high'] = data[
            'high'] if data['high'] > result['high'] else result['high']
        result['low'] = data['low'] if data['low'] < result['low'] else result[
            'low']
        result['volume_sold'] += data['volumeto']

    file_path = './out/' + fsym + '_crypto.xlsx'
    if not os.path.isfile(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append([
            'price', 'high', 'low', 'volume_sold', 'coin_supply', 'market_cap',
            'velocity', 'amount_hour', 'amount_avg', 'amount_spike',
            'sentiment_hour', 'sentiment_avg', 'sentiment_spike'
        ])
        wb.save(file_path)

    else:
        wb = load_workbook(file_path)
        ws = wb.active
    ws.append([
        time.strftime("%Y-%m-%d %H:%M:%S",
                      time.localtime()), result['price'], result['high'],
        result['low'], result['volume_sold'], result['coin_supply'],
        result['market_cap'], result['velocity'], result['amount_hour'],
        result['amount_avg'], result['amount_spike'], result['sentiment_hour'],
        result['sentiment_avg'], result['sentiment_spike']
    ])
    wb.save(file_path)
    return result


if __name__ == "__main__":
    print(gen_crypto_data('btc'))
